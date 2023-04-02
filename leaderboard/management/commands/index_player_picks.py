from django.core.management.base import BaseCommand
from logging import getLogger
from psycopg2.errors import UniqueViolation

from leaderboard.models import League, OverUnderLine, Pick, Player, PlayerScore, Season, Team

import openpyxl

LOG = getLogger(__name__)

class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str)

    def handle(self, *args, **options):
        file_name = options['file']
        if file_name:
            wb = openpyxl.load_workbook(file_name)

            if 'Picks' in wb.sheetnames:
                player_picks = self._read_player_picks_from_excel(wb['Picks'])
                self._bulk_create_player_picks_from_excel(player_picks)

            if 'Lines' in wb.sheetnames:
                over_under_lines = self._read_over_under_line_from_excel(wb['Lines'])
                self._bulk_create_over_under_lines(over_under_lines)


    def _read_player_picks_from_excel(self, sheet):
        player_picks = []

        for row in range(2, sheet.max_row + 1):
            player_name = str(sheet['A' + str(row)].value)
            pick1 = str(sheet['B' + str(row)].value)
            over1 = True if str(sheet['C' + str(row)].value) == 'O' else False
            pick2 = str(sheet['D' + str(row)].value)
            over2 = True if str(sheet['E' + str(row)].value) == 'O' else False
            pick3 = str(sheet['F' + str(row)].value)
            over3 = True if str(sheet['G' + str(row)].value) == 'O' else False
            pick4 = str(sheet['H' + str(row)].value)
            over4 = True if str(sheet['I' + str(row)].value) == 'O' else False
            league_name = str(sheet['J' + str(row)].value)
            season_year = str(sheet['K' + str(row)].value)

            league = League.objects.get(name=league_name)
            season = Season.objects.get(name=season_year, league=league)
            teams_dict = {team.abbreviation: team for team in Team.objects.filter(league=league)}

            player, created = Player.objects.get_or_create(name=player_name)
            PlayerScore.objects.get_or_create(player=player, season=season)

            player_picks = player_picks + [
                Pick(player=player, season=season, team=teams_dict[pick1], over=over1),
                Pick(player=player, season=season, team=teams_dict[pick2], over=over2),
                Pick(player=player, season=season, team=teams_dict[pick3], over=over3),
                Pick(player=player, season=season, team=teams_dict[pick4], over=over4),
            ]

        return player_picks

    def _bulk_create_player_picks_from_excel(self, player_picks):
        try:
            Pick.objects.bulk_create(player_picks)
        except UniqueViolation:
            print("Picks already imported.")
        except:
            print("Unknown error with importing picks.")

    def _read_over_under_line_from_excel(self, sheet):
        over_under_lines = []
        for row in range(2, sheet.max_row + 1):
            team_abbr = str(sheet['A' + str(row)].value)
            value = str(sheet['B' + str(row)].value)
            league_name = str(sheet['C' + str(row)].value)
            season_year = str(sheet['D' + str(row)].value)

            if League.objects.filter(name=league_name).count() > 0:
                league = League.objects.get(name=league_name)
                if Season.objects.filter(name=season_year, league=league).count() > 0:
                    season = Season.objects.get(name=season_year, league=league)
                    team = Team.objects.get(abbreviation=team_abbr, league=league)
                    over_under_lines.append(OverUnderLine(team=team, line=value, season=season))

        print(over_under_lines)

        return over_under_lines

    def _bulk_create_over_under_lines(self, over_under_lines):
        try:
            OverUnderLine.objects.bulk_create(over_under_lines)
        except UniqueViolation:
            print("Over/Under line already imported.")
        except:
            print("Unknown error with importing over/under line.")

    #def _generate_key_name_output_workbook(self, keys_dict, file_name):
        #wb = openpyxl.Workbook()
        #sheet = wb['Sheet']

        #sheet['A1'] = 'name'
        #sheet['B1'] = 'id'
        #i = 2
        #for key_dict in keys_dict:
            #sheet['A' + str(i)] = key_dict['name']
            #sheet['B' + str(i)] = key_dict['id']
            #i += 1

        #wb.save(file_name)